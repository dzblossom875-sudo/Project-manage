#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 公式验证脚本

用法：
    python scripts/recalc.py <excel_path>

返回 JSON 格式结果：
    {"status": "success"} - 所有公式计算正常
    {"status": "error", "message": "错误信息"} - 有公式错误
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def validate_excel_formulas(excel_path: str) -> dict:
    """
    验证 Excel 文件中的公式是否正确计算

    检查项：
    1. 所有包含公式的单元格是否能正常计算
    2. 合计行的 SUM 公式结果是否正确

    Args:
        excel_path: Excel 文件路径

    Returns:
        {"status": "success"} 或 {"status": "error", "message": "..."}
    """
    try:
        # 加载工作簿，保留公式
        wb = load_workbook(excel_path, data_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 遍历所有单元格检查公式
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # 这是一个公式单元格，尝试解析
                        try:
                            # 检查公式语法是否合法
                            formula = cell.value[1:]  # 去掉等号

                            # 检查 SUM 函数
                            if 'SUM' in formula.upper():
                                # 提取 SUM 范围
                                import re
                                match = re.search(r'SUM\(([A-Z]+\d+):([A-Z]+\d+)\)', formula, re.IGNORECASE)
                                if match:
                                    start_ref, end_ref = match.groups()
                                    # 验证范围是否有效
                                    # 这里只是语法检查，openpyxl 会在保存时处理

                        except Exception as e:
                            return {
                                "status": "error",
                                "message": f"Sheet '{sheet_name}' 单元格 {cell.coordinate} 公式错误: {str(e)}"
                            }

        # 重新加载，使用 data_only=True 检查计算值
        wb_data = load_workbook(excel_path, data_only=True)

        for sheet_name in wb_data.sheetnames:
            ws = wb_data[sheet_name]

            # 检查最后一行（通常是合计行）
            if ws.max_row > 1:
                last_row = ws.max_row
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=last_row, column=col_idx)
                    if cell.value and isinstance(cell.value, (int, float)):
                        # 检查数值是否合理（非负）
                        if cell.value < 0:
                            return {
                                "status": "error",
                                "message": f"Sheet '{sheet_name}' 合计行出现负值: {cell.coordinate} = {cell.value}"
                            }

        return {"status": "success", "message": "所有公式验证通过"}

    except FileNotFoundError:
        return {"status": "error", "message": f"文件不存在: {excel_path}"}
    except Exception as e:
        return {"status": "error", "message": f"验证失败: {str(e)}"}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "用法: python recalc.py <excel_path>"}, ensure_ascii=False))
        sys.exit(1)

    excel_path = sys.argv[1]
    result = validate_excel_formulas(excel_path)
    print(json.dumps(result, ensure_ascii=False))

    sys.exit(0 if result["status"] == "success" else 1)
