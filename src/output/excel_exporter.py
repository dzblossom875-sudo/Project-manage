"""
Excel 导出模块

职责：将合并后的数据导出为提交版 Excel，格式参考模板文件。
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List


# 样式常量
FONT_NAME = "微软雅黑"
FONT_SIZE = 10
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
HEADER_HEIGHT = 20

# 颜色
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 新增
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 状态变更
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# 列宽限制
MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 50


def _calc_col_width(values: List) -> int:
    """
    计算列宽（中文字符按2计算）

    Args:
        values: 列中的所有值

    Returns:
        计算后的列宽（已限制在 [MIN_COL_WIDTH, MAX_COL_WIDTH]）
    """
    max_len = 0
    for v in values:
        s = str(v) if v is not None else ""
        # 中文字符按2计算宽度
        width = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in s)
        max_len = max(max_len, width)

    # 加一些边距
    width = max_len + 2
    return max(MIN_COL_WIDTH, min(width, MAX_COL_WIDTH))


def _apply_header_style(ws, start_row: int = 1):
    """
    应用表头样式

    Args:
        ws: 工作表对象
        start_row: 表头所在行号（默认第1行）
    """
    for cell in ws[start_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 设置行高
    ws.row_dimensions[start_row].height = HEADER_HEIGHT


def _apply_data_style(ws, start_row: int = 2, change_type_col: int = None):
    """
    应用数据行样式

    Args:
        ws: 工作表对象
        start_row: 数据起始行号
        change_type_col: 变更类型列的索引（从1开始），用于判断行颜色
    """
    normal_font = Font(name=FONT_NAME, size=FONT_SIZE)
    wrap_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), start=start_row):
        # 判断行颜色
        fill = WHITE_FILL
        if change_type_col:
            change_type = row[change_type_col - 1].value
            if change_type == "新增":
                fill = GREEN_FILL
            elif change_type == "状态变更":
                fill = YELLOW_FILL

        for cell in row:
            cell.font = normal_font
            cell.fill = fill
            cell.alignment = wrap_alignment


def _setup_sheet(ws, df: pd.DataFrame, with_auto_filter: bool = True, with_freeze: bool = True):
    """
    设置工作表的基本格式

    Args:
        ws: 工作表对象
        df: 数据 DataFrame
        with_auto_filter: 是否开启自动筛选
        with_freeze: 是否冻结首行
    """
    if df.empty:
        return

    # 计算列宽
    for col_idx, col_name in enumerate(df.columns, 1):
        values = [col_name] + df[col_name].astype(str).tolist()
        width = _calc_col_width(values)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # 自动筛选
    if with_auto_filter and not df.empty:
        ws.auto_filter.ref = ws.dimensions

    # 冻结首行
    if with_freeze:
        ws.freeze_panes = "A2"


def _find_col_index(df: pd.DataFrame, col_name: str) -> int:
    """
    查找列的索引（从1开始，用于 openpyxl）

    Args:
        df: DataFrame
        col_name: 列名

    Returns:
        列索引（从1开始），找不到返回 None
    """
    try:
        return df.columns.get_loc(col_name) + 1
    except KeyError:
        return None


def export_excel(merged_df: pd.DataFrame, output_path: str) -> None:
    """
    导出提交版 Excel，格式参考 data/input/机构间REITs项目整理260309.xlsx

    包含4个 Sheet：
    - Sheet1「全部项目」：全部数据，按"交易所系统更新日期"降序，含颜色标注和合计行
    - Sheet2「新增项目」：仅新增，全表绿色
    - Sheet3「状态变更」：仅状态变更，全表黄色
    - Sheet4「跟进状态汇总」：按"跟进和投放状态"分组统计

    Args:
        merged_df: 合并后的 DataFrame（由 merger.merge 返回）
        output_path: 输出文件路径
    """
    if merged_df.empty:
        print("警告: 数据为空，无法导出 Excel")
        return

    # 确保变更类型列存在
    if "变更类型" not in merged_df.columns:
        merged_df["变更类型"] = "无变化"

    # 按"交易所系统更新日期"降序排序
    df_sorted = merged_df.sort_values(
        by="交易所系统更新日期",
        ascending=False,
        na_position="last"
    ).reset_index(drop=True)

    # 创建工作簿
    wb = Workbook()

    # ===== Sheet1: 全部项目 =====
    ws_all = wb.active
    ws_all.title = "全部项目"

    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df_sorted, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_all.cell(row=r_idx, column=c_idx, value=value)

    # 应用样式
    change_type_col = _find_col_index(df_sorted, "变更类型")
    _apply_header_style(ws_all, start_row=1)
    _apply_data_style(ws_all, start_row=2, change_type_col=change_type_col)
    _setup_sheet(ws_all, df_sorted)

    # 备注和变更详情列开启自动换行
    for col_name in ["备注", "变更详情"]:
        col_idx = _find_col_index(df_sorted, col_name)
        if col_idx:
            for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 添加合计行（最后一行）
    total_row = ws_all.max_row + 1
    amount_col = _find_col_index(df_sorted, "拟发行金额(亿元)")

    # 在 A 列写入"合计"
    ws_all.cell(row=total_row, column=1, value="合计")
    ws_all.cell(row=total_row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    ws_all.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # 拟发行金额列使用 SUM 公式
    if amount_col:
        start_cell = f"{get_column_letter(amount_col)}2"
        end_cell = f"{get_column_letter(amount_col)}{total_row - 1}"
        ws_all.cell(row=total_row, column=amount_col, value=f"=SUM({start_cell}:{end_cell})")
        ws_all.cell(row=total_row, column=amount_col).font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        ws_all.cell(row=total_row, column=amount_col).alignment = Alignment(horizontal="right", vertical="center")

    # ===== Sheet2: 新增项目 =====
    new_df = df_sorted[df_sorted["变更类型"] == "新增"].copy()
    if not new_df.empty:
        ws_new = wb.create_sheet("新增项目")
        for r_idx, row in enumerate(dataframe_to_rows(new_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_new.cell(row=r_idx, column=c_idx, value=value)

        _apply_header_style(ws_new, start_row=1)
        _setup_sheet(ws_new, new_df)

        # 全表绿色底色
        for row in ws_new.iter_rows(min_row=2, max_row=ws_new.max_row):
            for cell in row:
                cell.fill = GREEN_FILL
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)

    # ===== Sheet3: 状态变更 =====
    changed_df = df_sorted[df_sorted["变更类型"] == "状态变更"].copy()
    if not changed_df.empty:
        ws_chg = wb.create_sheet("状态变更")
        for r_idx, row in enumerate(dataframe_to_rows(changed_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_chg.cell(row=r_idx, column=c_idx, value=value)

        _apply_header_style(ws_chg, start_row=1)
        _setup_sheet(ws_chg, changed_df)

        # 全表黄色底色
        for row in ws_chg.iter_rows(min_row=2, max_row=ws_chg.max_row):
            for cell in row:
                cell.fill = YELLOW_FILL
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)

    # ===== Sheet4: 跟进状态汇总 =====
    ws_summary = wb.create_sheet("跟进状态汇总")

    # 按"跟进和投放状态"分组统计
    summary_data = []
    follow_col = "跟进和投放状态"
    amount_col_name = "拟发行金额(亿元)"

    if follow_col in df_sorted.columns and amount_col_name in df_sorted.columns:
        # 确保金额列为数值类型
        df_summary = df_sorted.copy()
        df_summary[amount_col_name] = pd.to_numeric(df_summary[amount_col_name], errors='coerce').fillna(0)

        # 分组统计
        groups = df_summary.groupby(follow_col, sort=False).agg({
            "债券名称": "count",
            amount_col_name: "sum"
        }).reset_index()
        groups.columns = ["跟进和投放状态", "项目数量", "拟发行金额合计(亿元)"]

        # 添加总计行
        total_projects = len(df_summary)
        total_amount = df_summary[amount_col_name].sum()

        # 写入表头
        headers = ["跟进和投放状态", "项目数量", "拟发行金额合计(亿元)"]
        for c_idx, header in enumerate(headers, 1):
            ws_summary.cell(row=1, column=c_idx, value=header)

        _apply_header_style(ws_summary, start_row=1)

        # 写入数据
        for r_idx, (_, row) in enumerate(groups.iterrows(), 2):
            ws_summary.cell(row=r_idx, column=1, value=row["跟进和投放状态"] if row["跟进和投放状态"] else "（未填写）")
            ws_summary.cell(row=r_idx, column=2, value=row["项目数量"])
            ws_summary.cell(row=r_idx, column=3, value=row["拟发行金额合计(亿元)"])

        # 写入总计行
        total_row_idx = len(groups) + 2
        ws_summary.cell(row=total_row_idx, column=1, value="合计")
        ws_summary.cell(row=total_row_idx, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        ws_summary.cell(row=total_row_idx, column=2, value=total_projects)
        ws_summary.cell(row=total_row_idx, column=2).font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
        ws_summary.cell(row=total_row_idx, column=3, value=total_amount)
        ws_summary.cell(row=total_row_idx, column=3).font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)

        # 设置列宽
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 12
        ws_summary.column_dimensions["C"].width = 20

        # 数据样式
        for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
            for cell in row:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
                cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "right", vertical="center")

    # 保存文件
    wb.save(output_path)
    print(f"Excel 已导出: {output_path}")
