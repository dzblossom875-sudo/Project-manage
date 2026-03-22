"""
数据汇总、基准合并和Excel导出模块

DEPRECATED: 此模块已废弃，请使用 merger.py + output/excel_exporter.py 代替
此文件保留以兼容旧版调用，新代码请使用新的输出层架构

核心逻辑:
  以基准文件为底，合并爬取的新数据:
  - 基准中已有 + 爬取中也有 → 用爬取数据更新字段，标注"状态变更"或"无变化"
  - 基准中已有 + 爬取中没有 → 保留基准原样，标注"无变化"
  - 基准中没有 + 爬取中新增 → 追加，标注"新增"

新版调用方式:
    from merger import merge
    from output.excel_exporter import export_excel
    from output.html_exporter import export_html

    merged_df = merge(sse_projects, szse_projects, baseline_df)
    export_excel(merged_df, excel_path)
    export_html(merged_df, html_path)
"""

import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
from dataclasses import asdict

from baseline_reader import STANDARD_COLUMNS, COMPARE_FIELDS, make_key


class DataExporter:

    def __init__(self, output_dir: str = "data/output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def crawled_to_df(self, sse_projects: List, szse_projects: List) -> pd.DataFrame:
        """将爬取的项目对象列表转为标准化 DataFrame"""
        all_projects = sse_projects + szse_projects
        if not all_projects:
            return pd.DataFrame()

        data = [asdict(p) for p in all_projects]
        df = pd.DataFrame(data)

        column_mapping = {
            "exchange": "交易所",
            "bond_name": "债券名称",
            "manager": "计划管理人",
            "bond_type": "品种",
            "amount": "拟发行金额(亿元)",
            "status": "项目状态",
            "accept_date": "受理日期",
            "update_date": "更新日期",
        }
        df = df.rename(columns=column_mapping)
        return df

    def merge_with_baseline(
        self,
        crawled_df: pd.DataFrame,
        baseline_df: pd.DataFrame,
        baseline_index: Dict[str, dict],
    ) -> pd.DataFrame:
        """
        以基准文件为底，合并爬取数据，标注变更

        Returns:
            合并后的完整 DataFrame，包含 变更类型 和 变更详情 列
        """
        crawled_index = {}
        for _, row in crawled_df.iterrows():
            ex = str(row.get("交易所", "")).strip()
            bn = str(row.get("债券名称", "")).strip()
            if bn:
                crawled_index[make_key(ex, bn)] = row.to_dict()

        rows_out = []

        for _, base_row in baseline_df.iterrows():
            ex = str(base_row.get("交易所", "")).strip()
            bn = str(base_row.get("债券名称", "")).strip()
            key = make_key(ex, bn)

            row_dict = base_row.to_dict()

            if key in crawled_index:
                crawled_row = crawled_index.pop(key)
                changes = self._detect_field_changes(row_dict, crawled_row)

                if changes:
                    for field in COMPARE_FIELDS:
                        new_val = str(crawled_row.get(field, "")).strip()
                        if new_val and new_val != "nan":
                            row_dict[field] = new_val
                    row_dict["变更类型"] = "状态变更"
                    row_dict["变更详情"] = "; ".join(changes)
                else:
                    row_dict["变更类型"] = "无变化"
                    row_dict["变更详情"] = ""
            else:
                row_dict["变更类型"] = "无变化"
                row_dict["变更详情"] = ""

            rows_out.append(row_dict)

        for key, crawled_row in crawled_index.items():
            row_dict = crawled_row.copy()
            row_dict["变更类型"] = "新增"
            row_dict["变更详情"] = "基准文件中不存在"
            rows_out.append(row_dict)

        result = pd.DataFrame(rows_out)

        col_order = [c for c in STANDARD_COLUMNS if c in result.columns]
        extra = [c for c in result.columns if c not in col_order and c not in ("变更类型", "变更详情")]
        col_order = col_order + extra + ["变更类型", "变更详情"]
        col_order = [c for c in col_order if c in result.columns]
        result = result[col_order]

        new_count = sum(1 for t in result["变更类型"] if t == "新增")
        changed_count = sum(1 for t in result["变更类型"] if t == "状态变更")
        unchanged_count = sum(1 for t in result["变更类型"] if t == "无变化")

        print(f"\n=== 合并统计 ===")
        print(f"基准项目: {len(baseline_df)} 条")
        print(f"本次爬取: {len(crawled_df)} 条")
        print(f"合并后:   {len(result)} 条")
        print(f"  - 新增:     {new_count}")
        print(f"  - 状态变更: {changed_count}")
        print(f"  - 无变化:   {unchanged_count}")
        print(f"================\n")

        return result

    def merge_crawled_only(
        self,
        crawled_df: pd.DataFrame,
    ) -> pd.DataFrame:
        """没有基准文件时，直接输出爬取数据"""
        if crawled_df.empty:
            return crawled_df
        crawled_df = crawled_df.copy()
        crawled_df["变更类型"] = "新增"
        crawled_df["变更详情"] = ""
        return crawled_df

    def _detect_field_changes(self, old: dict, new: dict) -> List[str]:
        """比较两行数据的关键字段，返回变更描述列表"""
        changes = []
        for field in COMPARE_FIELDS:
            old_val = str(old.get(field, "")).strip()
            new_val = str(new.get(field, "")).strip()
            if old_val == "nan" or old_val == "NaT":
                old_val = ""
            if new_val == "nan" or new_val == "NaT":
                new_val = ""
            if old_val != new_val and new_val:
                changes.append(f"{field}: {old_val} -> {new_val}")
        return changes

    def export_to_excel(self, df: pd.DataFrame, filename: str = None) -> str:
        """导出数据到Excel, 包含多个sheet"""
        if df.empty:
            print("警告: 数据为空，无法导出")
            return ""

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"持有型ABS项目汇总_{timestamp}.xlsx"

        output_path = os.path.join(self.output_dir, filename)

        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="全部项目", index=False)

                if "变更类型" in df.columns:
                    new_df = df[df["变更类型"] == "新增"]
                    if not new_df.empty:
                        new_df.to_excel(writer, sheet_name="新增项目", index=False)

                    changed_df = df[df["变更类型"] == "状态变更"]
                    if not changed_df.empty:
                        changed_df.to_excel(writer, sheet_name="状态变更", index=False)

                for sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    for col in ws.columns:
                        max_len = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_len = max(max_len, len(str(cell.value)))
                            except Exception:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

                    self._highlight_changes(ws)

            print(f"数据已导出到: {output_path}")
            print(f"共 {len(df)} 条记录")
            return output_path

        except Exception as e:
            print(f"导出Excel失败: {e}")
            return ""

    def _highlight_changes(self, ws):
        """新增行绿色, 状态变更行黄色"""
        try:
            from openpyxl.styles import PatternFill

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            header_row = [cell.value for cell in ws[1]]
            if "变更类型" not in header_row:
                return

            change_col_idx = header_row.index("变更类型") + 1

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                change_type = row[change_col_idx - 1].value
                if change_type == "新增":
                    for cell in row:
                        cell.fill = green_fill
                elif change_type == "状态变更":
                    for cell in row:
                        cell.fill = yellow_fill
        except Exception:
            pass


def export_projects(
    sse_projects: List,
    szse_projects: List,
    output_dir: str = "data/output",
    baseline_df: pd.DataFrame = None,
    baseline_index: Dict[str, dict] = None,
) -> str:
    """
    便捷函数: 基准+爬取合并后导出Excel

    - 有基准: 以基准为底，合并爬取数据，标注变更
    - 无基准: 仅输出爬取数据
    """
    exporter = DataExporter(output_dir)
    crawled_df = exporter.crawled_to_df(sse_projects, szse_projects)

    print(f"\n=== 爬取汇总 ===")
    print(f"上交所: {len(sse_projects)} 条")
    print(f"深交所: {len(szse_projects)} 条")
    print(f"================")

    if baseline_df is not None and baseline_index:
        merged = exporter.merge_with_baseline(crawled_df, baseline_df, baseline_index)
    elif not crawled_df.empty:
        merged = exporter.merge_crawled_only(crawled_df)
    else:
        print("没有任何数据可导出")
        return ""

    return exporter.export_to_excel(merged)
